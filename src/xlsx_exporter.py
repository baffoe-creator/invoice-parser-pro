import os
from typing import Dict, Any, List
from datetime import datetime
import uuid
import logging

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl not available - Excel export will be limited")

logger = logging.getLogger(__name__)


class XLSXExporter:
    def __init__(self, session_id: str = None):
        IS_VERCEL = os.getenv("VERCEL") == "1"
        self.data_dir = "/tmp/data" if IS_VERCEL else "data"

        self.session_id = session_id or f"session_{uuid.uuid4().hex[:8]}"
        self.xlsx_file_path = f"{self.data_dir}/parsed_invoices_{self.session_id}.xlsx"
        self._ensure_data_directory()
        self.expected_columns = [
            "file_name",
            "vendor",
            "invoice_number",
            "invoice_date",
            "subtotal",
            "shipping_amount",
            "tax_amount",
            "total_amount",
            "currency",
            "line_item_description",
            "line_item_quantity",
            "line_item_unit_price",
            "line_item_amount",
            "parsed_timestamp",
        ]

    def _ensure_data_directory(self):
        try:
            os.makedirs(self.data_dir, exist_ok=True)
        except Exception as e:
            logger.warning(f"Could not create directory {self.data_dir}: {e}")
            if os.getenv("VERCEL"):
                raise

    def _create_new_workbook(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required but not available")

        wb = Workbook()

        ws_main = wb.active
        ws_main.title = "Invoice Data"
        ws_main.append(self.expected_columns)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        for cell in ws_main[1]:
            cell.font = header_font
            cell.fill = header_fill

        ws_meta = wb.create_sheet("Metadata")
        ws_meta.append(["Session ID", "Created", "Total Invoices", "Last Updated"])
        ws_meta.append(
            [self.session_id, datetime.now().isoformat(), 0, datetime.now().isoformat()]
        )

        ws_cols = wb.create_sheet("Column Definitions")
        ws_cols.append(["Column Name", "Data Type", "Description"])
        column_defs = [
            ["file_name", "Text", "Original PDF filename"],
            ["vendor", "Text", "Vendor/company name"],
            ["invoice_number", "Text", "Invoice number"],
            ["invoice_date", "Text", "Invoice date"],
            ["subtotal", "Number", "Subtotal amount before discounts/tax"],
            ["shipping_amount", "Number", "Shipping cost"],
            ["tax_amount", "Number", "Tax amount"],
            ["total_amount", "Number", "Total amount due"],
            ["currency", "Text", "Currency code (USD, EUR, etc)"],
            ["line_item_description", "Text", "Product/service description"],
            ["line_item_quantity", "Number", "Quantity of items"],
            ["line_item_unit_price", "Number", "Unit price per item"],
            ["line_item_amount", "Number", "Line total (quantity × unit price)"],
            ["parsed_timestamp", "Text", "When invoice was parsed"],
        ]
        for defn in column_defs:
            ws_cols.append(defn)

        return wb

    def _normalize_data(self, normalized_data: Dict[str, Any]) -> Dict[str, Any]:
        normalized = {}

        for column in self.expected_columns:
            value = normalized_data.get(column)

            if column in [
                "subtotal",
                "shipping_amount",
                "tax_amount",
                "total_amount",
                "line_item_quantity",
                "line_item_unit_price",
                "line_item_amount",
            ]:
                try:
                    normalized[column] = float(value) if value is not None else 0.0
                except (ValueError, TypeError):
                    normalized[column] = 0.0
            else:
                if value is None:
                    normalized[column] = ""
                else:
                    normalized[column] = str(value).strip()

        return normalized

    def append_normalized_data(
        self, normalized_data: Dict[str, Any], filename: str
    ) -> Dict[str, Any]:
        if not OPENPYXL_AVAILABLE:
            return {
                "success": False,
                "error": "openpyxl not available - cannot export to Excel",
                "filename": filename,
            }

        try:
            logger.info(f"Appending normalized data to XLSX for: {filename}")

            clean_data = self._normalize_data(normalized_data)

            if os.path.exists(self.xlsx_file_path):
                try:
                    existing_df = pd.read_excel(
                        self.xlsx_file_path, sheet_name="Invoice Data"
                    )
                except Exception as e:
                    logger.warning(
                        f"Could not read existing file, creating new: {str(e)}"
                    )
                    existing_df = pd.DataFrame(columns=self.expected_columns)
            else:
                existing_df = pd.DataFrame(columns=self.expected_columns)

            new_row = {col: clean_data.get(col, "") for col in self.expected_columns}
            new_df = pd.DataFrame([new_row])

            combined_df = pd.concat([existing_df, new_df], ignore_index=True)

            with pd.ExcelWriter(self.xlsx_file_path, engine="openpyxl") as writer:
                combined_df.to_excel(writer, sheet_name="Invoice Data", index=False)

                if not combined_df.empty and "invoice_date" in combined_df.columns:
                    date_series = pd.to_datetime(
                        combined_df["invoice_date"], errors="coerce"
                    )
                    min_date = date_series.min()
                    max_date = date_series.max()

                    earliest_date = (
                        min_date.strftime("%b %d %Y")
                        if not pd.isna(min_date)
                        else "N/A"
                    )
                    latest_date = (
                        max_date.strftime("%b %d %Y")
                        if not pd.isna(max_date)
                        else "N/A"
                    )
                else:
                    earliest_date = "N/A"
                    latest_date = "N/A"

                stats_data = {
                    "Statistic": [
                        "Session ID",
                        "Total Invoices",
                        "Earliest Date",
                        "Latest Date",
                        "Total Amount",
                        "Last Updated",
                    ],
                    "Value": [
                        self.session_id,
                        len(combined_df),
                        earliest_date,
                        latest_date,
                        (
                            combined_df["total_amount"].sum()
                            if not combined_df.empty
                            and "total_amount" in combined_df.columns
                            else 0
                        ),
                        datetime.now().isoformat(),
                    ],
                }
                pd.DataFrame(stats_data).to_excel(
                    writer, sheet_name="Statistics", index=False
                )

                column_defs = pd.DataFrame(
                    [
                        ["file_name", "Text", "Original PDF filename"],
                        ["vendor", "Text", "Vendor/company name"],
                        ["invoice_number", "Text", "Invoice number"],
                        ["invoice_date", "Text", "Invoice date"],
                        ["subtotal", "Number", "Subtotal amount before discounts/tax"],
                        ["shipping_amount", "Number", "Shipping cost"],
                        ["tax_amount", "Number", "Tax amount"],
                        ["total_amount", "Number", "Total amount due"],
                        ["currency", "Text", "Currency code (USD, EUR, etc)"],
                        [
                            "line_item_description",
                            "Text",
                            "Product/service description",
                        ],
                        ["line_item_quantity", "Number", "Quantity of items"],
                        ["line_item_unit_price", "Number", "Unit price per item"],
                        [
                            "line_item_amount",
                            "Number",
                            "Line total (quantity × unit price)",
                        ],
                        ["parsed_timestamp", "Text", "When invoice was parsed"],
                    ],
                    columns=["Column Name", "Data Type", "Description"],
                )
                column_defs.to_excel(
                    writer, sheet_name="Column Definitions", index=False
                )

            logger.info(f"Successfully appended data for: {filename}")

            return {
                "success": True,
                "message": "Data appended to XLSX file",
                "filename": filename,
                "file_path": self.xlsx_file_path,
                "session_id": self.session_id,
            }

        except Exception as e:
            logger.error(f"XLSX append failed for {filename}: {str(e)}", exc_info=True)
            return {
                "success": False,
                "error": f"XLSX append failed: {str(e)}",
                "filename": filename,
            }

    def get_file_data(self) -> bytes:
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl not available - cannot read Excel files")

        try:
            with open(self.xlsx_file_path, "rb") as f:
                return f.read()
        except Exception as e:
            logger.error(f"Error reading XLSX file: {str(e)}")
            raise

    def get_file_stats(self) -> Dict[str, Any]:
        if not OPENPYXL_AVAILABLE:
            return {
                "exists": False,
                "error": "openpyxl not available",
                "row_count": 0,
                "session_id": self.session_id,
            }

        if not os.path.exists(self.xlsx_file_path):
            return {"exists": False, "row_count": 0, "session_id": self.session_id}

        try:
            df = pd.read_excel(self.xlsx_file_path, sheet_name="Invoice Data")

            total_amount = 0
            if "total_amount" in df.columns and not df.empty:
                total_amount = float(df["total_amount"].sum())

            return {
                "exists": True,
                "file_path": self.xlsx_file_path,
                "session_id": self.session_id,
                "row_count": int(len(df)),
                "columns": [str(col) for col in df.columns] if not df.empty else [],
                "total_amount": total_amount,
                "file_size": int(os.path.getsize(self.xlsx_file_path)),
            }
        except Exception as e:
            logger.error(f"Error getting XLSX stats: {str(e)}")
            return {"exists": False, "error": str(e), "session_id": self.session_id}

    def create_new_file(self) -> Dict[str, Any]:
        if not OPENPYXL_AVAILABLE:
            return {
                "success": False,
                "error": "openpyxl not available - cannot create Excel files",
            }

        try:
            new_session_id = f"session_{uuid.uuid4().hex[:8]}"
            old_path = self.xlsx_file_path
            self.session_id = new_session_id
            self.xlsx_file_path = (
                f"{self.data_dir}/parsed_invoices_{new_session_id}.xlsx"
            )

            wb = self._create_new_workbook()
            wb.save(self.xlsx_file_path)

            return {
                "success": True,
                "message": "New XLSX file created",
                "file_path": self.xlsx_file_path,
                "session_id": new_session_id,
                "old_path": old_path,
            }
        except Exception as e:
            logger.error(f"Failed to create new XLSX: {str(e)}")
            return {"success": False, "error": f"Failed to create new XLSX: {str(e)}"}

    def get_all_sessions(self) -> List[Dict[str, Any]]:
        if not OPENPYXL_AVAILABLE:
            return []

        sessions = []
        if os.path.exists(self.data_dir):
            for file in os.listdir(self.data_dir):
                if file.startswith("parsed_invoices_session_") and file.endswith(
                    ".xlsx"
                ):
                    session_id = file.replace("parsed_invoices_", "").replace(
                        ".xlsx", ""
                    )
                    file_path = os.path.join(self.data_dir, file)
                    try:
                        df = pd.read_excel(file_path, sheet_name="Invoice Data")
                        sessions.append(
                            {
                                "session_id": session_id,
                                "file_path": file_path,
                                "row_count": len(df),
                                "created_time": datetime.fromtimestamp(
                                    os.path.getctime(file_path)
                                ).isoformat(),
                            }
                        )
                    except:
                        continue
        return sessions
